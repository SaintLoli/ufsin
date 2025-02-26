import openpyxl
from io import BytesIO
from classes.DB_helper import DBHelper
from classes.TABLENAMES import TABLES

class Report:
    def __init__(self, template_name, database_path="../DB/UFSIN_DB.db", database="", template_path="../templates/"):
        if database:
            self.database = database
        else:
            self.database = DBHelper(filepath=database_path)

        if template_name == "items_on_warehouses_ok_status":
            self.workbook = openpyxl.load_workbook(template_path + "items_on_warehouses_ok_status.xlsx")
            self.fill_items_on_warehouses(status="ok")
        elif template_name == "items_on_warehouses_remove_status":
            self.workbook = openpyxl.load_workbook(template_path + "items_on_warehouses_remove_status.xlsx")
            self.fill_items_on_warehouses(status="removed")
        elif template_name == "items_on_warehouses_underchange_status":
            print(23233224323432)
            self.workbook = openpyxl.load_workbook(template_path + "items_on_warehouses_underchange_status.xlsx")
            self.fill_items_on_warehouses(status="underchange")
        else:
            self.workbook = openpyxl.load_workbook(template_path + "empty_template.xlsx")

    def fill_items_on_warehouses(self, status="ok"):
        sheet = self.workbook.active
        row, column = 2, 1
        cell = sheet.cell(row, column)

        for warehouse in self.database.get_warehouses():
            if status == "ok":
                all_warehouse_items = self.database.get_item_name_by_warehouse_ok_status(warehouse[0])
            elif status == "removed":
                all_warehouse_items = self.database.get_item_name_by_warehouse_remove_status(warehouse[0])
            else:
                all_warehouse_items = self.database.get_item_name_by_warehouse_underchange_status(warehouse[0])

            if len(all_warehouse_items) != 0:
                cell.value = f"{warehouse[0]}\n {warehouse[1]}"

                row += 1
                column += 1
                cell = sheet.cell(row, column)

                for GROUP_NAME in TABLES.keys():
                    typed_warehouse_items = [i for i in all_warehouse_items if i[0] == GROUP_NAME]
                    other_typed_items = [i for i in all_warehouse_items if i[0] == "other"]

                    if typed_warehouse_items:
                        if GROUP_NAME != 'other':
                            cell.value = TABLES[GROUP_NAME]

                            row += 1
                            column += 1
                            cell = sheet.cell(row, column)

                            for item in typed_warehouse_items:
                                cell.value = self.database.get_item_name_by_type(item[1], item[0])

                                column += 1
                                cell = sheet.cell(row, column)

                                cell.value = item[2]
                                row += 1
                                column -= 1
                                cell = sheet.cell(row, column)

                            column -= 1
                            cell = sheet.cell(row, column)
                        else:
                            for spec_type in self.database.get_other_types():
                                if spec_type[0] in set(self.database.get_item_type_by_id(i[1])[0] for i in other_typed_items):
                                    cell.value = spec_type[0]

                                    row += 1
                                    column += 1
                                    cell = sheet.cell(row, column)

                                    for spec_item in typed_warehouse_items:

                                        if spec_type[0] == self.database.get_item_type_by_id(spec_item[1])[0]:
                                            cell.value = self.database.get_item_name_by_type(spec_item[1], spec_item[0])

                                            column += 1
                                            cell = sheet.cell(row, column)

                                            cell.value = spec_item[2]
                                            row += 1
                                            column -= 1
                                            cell = sheet.cell(row, column)

                                    column -= 1
                                    cell = sheet.cell(row, column)

                column -= 1
                cell = sheet.cell(row, column)

        sheet.column_dimensions["A"].width = 45
        sheet.column_dimensions["B"].width = 25
        sheet.column_dimensions["C"].width = 25
        sheet.column_dimensions["D"].width = 10

        self.get_bytes_stream()
        self.workbook.save('new_example.xlsx')


    def get_bytes_stream(self):
        # Создаем поток байтов
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)  # Перемещаем указатель в начало
        return output


if __name__ == "__main__":
    report = Report("items_on_warehouses_underchange_status")